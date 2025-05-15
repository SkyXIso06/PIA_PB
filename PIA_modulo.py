import os
import re
import socket
from OTXv2 import OTXv2
import matplotlib.pyplot as plt
from collections import Counter
import json
import statistics
import pandas as pd


#Consulatr la API KEY

if os.path.exists("key.txt"):
    with open("key.txt", "r", encoding="utf-8") as f:
        api_key = f.read().strip()  
        otx = OTXv2(api_key)
else:
    print("El archivo key.txt no existe.")


def hay_conexion():
    try:
        socket.create_connection(("8.8.8.8", 53), timeout=5)
        return True
    except (socket.error, socket.timeout):
        return False

def obtener_indicadores_otx(pulse_id, limite):
    try:
        indicators = otx.get_pulse_indicators(pulse_id)
        return indicators[:limite]
    except Exception as e:
        print(f"Hubo un error al obtener los indicadores: {e}")
        return []

# -------------------------
# Manejo de archivos TXT
# -------------------------

def cargar_datos_txt(nombre_archivo):
    if os.path.exists(nombre_archivo):
        with open(nombre_archivo, "r", encoding="utf-8") as archivo:
            try:
                return json.load(archivo)  # Cargar como un objeto JSON completo
            except json.JSONDecodeError:
                return []
    return []

def guardar_datos_txt(nuevos_datos, nombre_archivo):
    datos_guardados = cargar_datos_txt(nombre_archivo)
    indicadores_existentes = {item["indicator"] for item in datos_guardados}

    nuevos_unicos = [
        {
            "type": item["type"],
            "indicator": item["indicator"],
            "created": item["created"],
            "pulse_key": item["pulse_key"],
            "id": item["id"]
        }
        for item in nuevos_datos
        if item["indicator"] not in indicadores_existentes
    ]

    datos_guardados.extend(nuevos_unicos)

    with open(nombre_archivo, "w", encoding="utf-8") as archivo:
        json.dump(datos_guardados, archivo, indent=4, ensure_ascii=False)

    print(f"Se han guardado correctamente {len(nuevos_unicos)} nuevos indicadores en {nombre_archivo}.")

# -------------------------
# Usuarios
# -------------------------

def registrar():
    while True:
        os.system("clear")
        try:
            print("-" * 32)
            print("1- Iniciar sesion\n2- Registrarse\n3- Salir")
            print("-" * 32)
            menu = int(input("\nOpcion: "))
            if menu == 1:
                valid = ingresar_usuario()
                if valid == True:
                    break
            elif menu == 2:
                registrar_usuario()
            elif menu == 3:
                input("Gracias por usar el programa.")
                os.system("close()")
            else:
                input("Opción no válida")
        except:
            input("Debes ingresar un número")

def ingresar_usuario():
    valid = False
    while True:
        os.system("clear")
        correo = input("Ingrese su correo electrónico: \n")
        contraseña = input("Ingrese su contraseña: \n")
        if verificar_usuario(correo, contraseña):
            input("Inicio de sesión exitoso.")
            valid = True
            break
        else:
            input("Correo o contraseña incorrectos. Intente nuevamente.")
            break
    return valid


def registrar_usuario():
    while True:
        os.system("clear")
        correo = input("Ingrese su correo electrónico de registro: \n")
        if not validar_correo(correo):
            input(
                "El correo no es válido. Debe seguir el formato usuario@mail.com. Intente nuevamente."
            )
        else:
            break

    while True:
        contraseña = input("Ingrese su contraseña de registro: \n")
        if not validar_contraseña(contraseña):
            input(
                "La contraseña debe contener al menos una mayúscula y un número. Intente nuevamente."
            )
        else:
            break

    if guardar_usuario(correo, contraseña):
        input("Usuario registrado exitosamente.")
    else:
        input("El correo ya está registrado. Intente con otro correo.")

def guardar_usuario(correo, contrasena):
    datos = []

    if os.path.exists("usuarios.txt"):
        with open("usuarios.txt", "r", encoding="utf-8") as archivo:
            try:
                datos = json.load(archivo)
            except:
                datos = []
        for usuario in datos:
            if usuario["correo"] == correo:
                return False
            
    datos.append({
        "correo": correo,
        "contrasena": contrasena
    })
    with open("usuarios.txt", "w", encoding="utf-8") as archivo:
        json.dump(datos, archivo, indent=4)
    return True

def verificar_usuario(correo, contrasena):
    if not os.path.exists("usuarios.txt"):
        return False

    with open("usuarios.txt", "r", encoding="utf-8") as archivo:
        try:
            datos = json.load(archivo)
        except:
            return False  
    for usuario in datos:
        if usuario["correo"] == correo and usuario["contrasena"] == contrasena:
            return True  
    return False


def validar_correo(correo):
    return re.match(r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$', correo)

def validar_contraseña(contrasena):
    return re.match(r'^(?=.*[A-Z])(?=.*\d).+$', contrasena)

# -------------------------
# Análisis y gráficas
# -------------------------

def contar_años_por_regex(datos):
    años = []
    for item in datos:
        match = re.search(r"\d{4}-\d{2}-\d{2}", item.get("created", ""))
        if match:
            año = match.group()[:4]
            años.append(año)
    return años

def contar_cantidad_por_año(años):
    return dict(Counter(años))

def graficar_frecuencia_por_año(conteo,nombre):
    if not conteo:
        print("No hay datos para graficar.")
        return

    años = sorted(conteo.keys(), key=lambda x: int(x))
    valores = [conteo[año] for año in años]

    valores = sorted(valores)

    # Calcular estadísticas
    media = statistics.mean(valores)
    mediana = statistics.median(valores)
    desviacion = statistics.stdev(valores) if len(valores) > 1 else 0.0
    moda = statistics.mode(valores) if len(set(valores)) != len(valores) else "No hay moda (valores únicos)"

    # Mostrar estadísticas en consola
    print("\nEstadísticas de frecuencia por año:")
    print(f"Media: {media:.2f}")
    print(f"Mediana: {mediana}")
    print(f"Desviación estándar: {desviacion:.2f}")
    print(f"Moda: {moda}")

    # Gráfica
    plt.figure(figsize=(10, 6))
    plt.bar(años, valores, color='blue')
    plt.xticks(rotation=45)
    plt.xlabel("Año de creación del indicador")
    plt.ylabel("Cantidad")
    plt.title("Indicadores maliciosos por año")
    plt.tight_layout()
    plt.grid(True, axis='y', linestyle='--', alpha=0.7)
    plt.show()
    estadisticas = {
    "Media": media,
    "Mediana": mediana,
    "Moda": moda,
    "Desviación estándar": desviacion
}
    exportar_estadisticas_excel(nombre, estadisticas)


def graficaiOs():
    archivo_txt = "iOs_maliciosos.txt"
    datos = cargar_datos_txt(archivo_txt)

    if not datos:
        print("No hay datos para graficar.")
        return

    tipos_validos = ['domain', 'hostname', 'IPv4', 'FileHash-SHA256', 'URL']
    tipos = [item['type'] for item in datos if item['type'] in tipos_validos]

    if tipos:
        conteo = Counter(tipos)
        cantidades = list(conteo.values())
        cantidades = sorted(cantidades)

        # Estadísticas
        moda = conteo.most_common(1)[0]  # (tipo, cantidad)
        media = statistics.mean(cantidades)
        mediana = statistics.median(cantidades)
        desviacion = statistics.stdev(cantidades) if len(cantidades) > 1 else 0.0

        # Mostrar en consola
        print("\nConteo de tipos de indicadores maliciosos para iOS:")
        for tipo, cantidad in conteo.items():
            print(f"{tipo}: {cantidad}")

        print("\nEstadísticas:")
        print(f"Moda: {moda[0]} ({moda[1]} veces)")
        print(f"Media: {media:.2f}")
        print(f"Mediana: {mediana}")
        print(f"Desviación estándar: {desviacion:.2f}")

        # Gráfico de pastel
        plt.figure(figsize=(8, 8))
        plt.pie(
            conteo.values(),
            labels=conteo.keys(),
            autopct='%1.1f%%',
            startangle=140,
            wedgeprops={'linewidth': 1, 'edgecolor': 'white'}
        )
        plt.title("Distribución de Tipos de Indicadores Maliciosos (iOS)")
        plt.axis('equal')
        plt.tight_layout()
        plt.show()
    else:
        print("No hay tipos válidos de indicadores maliciosos para graficar.")
    estadisticas = {
    "Media": media,
    "Mediana": mediana,
    "Moda": moda,
    "Desviación estándar": desviacion
}
    exportar_estadisticas_excel(archivo_txt, estadisticas)


def grafica_malware():
    archivo_txt = "malware.txt"
    datos = cargar_datos_txt(archivo_txt)

    if not datos:
        print("No hay datos para graficar.")
        return

    tipos_validos = ['domain', 'hostname','CIDR','IPv4','FileHash-SHA256','URL']
    tipos = [item['type'] for item in datos if item['type'] in tipos_validos]

    if tipos:
        conteo = Counter(tipos)
        cantidades = list(conteo.values())
        cantidades = sorted(cantidades)

        # Estadísticas
        moda = conteo.most_common(1)[0]  # (tipo, cantidad)
        media = statistics.mean(cantidades)
        mediana = statistics.median(cantidades)
        desviacion = statistics.stdev(cantidades) if len(cantidades) > 1 else 0.0

        # Mostrar en consola
        print("\nConteo de tipos de indicadores maliciosos :")
        for tipo, cantidad in conteo.items():
            print(f"{tipo}: {cantidad}")

        print("\nEstadísticas:")
        print(f"Moda: {moda[0]} ({moda[1]} veces)")
        print(f"Media: {media:.2f}")
        print(f"Mediana: {mediana}")
        print(f"Desviación estándar: {desviacion:.2f}")

        # Gráfico de pastel
        plt.figure(figsize=(8, 8))
        plt.pie(
            conteo.values(),
            labels=conteo.keys(),
            autopct='%1.1f%%',
            startangle=140,
            wedgeprops={'linewidth': 1, 'edgecolor': 'white'}
        )
        plt.title("Distribución de Tipos de Indicadores Maliciosos ")
        plt.axis('equal')
        plt.tight_layout()
        plt.show()
    else:
        print("No hay tipos válidos de indicadores maliciosos para graficar.")
    estadisticas = {
    "Media": media,
    "Mediana": mediana,
    "Moda": moda,
    "Desviación estándar": desviacion
}
    exportar_estadisticas_excel(archivo_txt, estadisticas)   

def graficar_dominios_maliciosos():
    archivo_txt = "dominios_maliciosos.txt"
    datos = cargar_datos_txt(archivo_txt)

    if datos:
        años = contar_años_por_regex(datos)
        conteo = contar_cantidad_por_año(años)
        graficar_frecuencia_por_año(conteo,archivo_txt)
    else:
        print("No hay datos disponibles para graficar.")

def graficar_links_maliciosos():
    archivo_txt = "google_links_maliciosos.txt"
    datos = cargar_datos_txt(archivo_txt)

    if datos:
        años = contar_años_por_regex(datos)
        conteo = contar_cantidad_por_año(años)
        graficar_frecuencia_por_año(conteo,archivo_txt)
    else:
        print("No hay datos disponibles para graficar.")
def verificarsistema(correo,contraseña):
     if os.path.exists("usuarios.txt"):
        with open("usuarios.txt", 'r', encoding='utf-8') as archivo:
            for linea in archivo:
                correo_guardado, contrasena_guardada = linea.strip().split(",")
                if correo==correo_guardado and contraseña==contrasena_guardada :
                    return True
     return False

def exportar_estadisticas_excel(nombre_archivo, estadisticas):
    from openpyxl import load_workbook

    ruta = f"estadisticas_{nombre_archivo.replace('.txt', '')}.xlsx"
    nombre_hoja = "Estadísticas"

    df_nuevo = pd.DataFrame([estadisticas])

    if os.path.exists(ruta):
        # Cargar el archivo y hoja existente
        libro = load_workbook(ruta)

        if nombre_hoja in libro.sheetnames:
            # Leer datos existentes
            df_existente = pd.read_excel(ruta, sheet_name=nombre_hoja)
            # Concatenar
            df_total = pd.concat([df_existente, df_nuevo], ignore_index=True)
        else:
            df_total = df_nuevo

        # Sobrescribir hoja completa
        with pd.ExcelWriter(ruta, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df_total.to_excel(writer, sheet_name=nombre_hoja, index=False)

    else:
        # Crear archivo nuevo
        with pd.ExcelWriter(ruta, engine="openpyxl") as writer:
            df_nuevo.to_excel(writer, sheet_name=nombre_hoja, index=False)

    print(f"Estadísticas exportadas a {ruta} (hoja: {nombre_hoja})")
